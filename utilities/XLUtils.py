import openpyxl

def getRowCount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)


def getColumnCount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_coulumn)


def readData(path, sheetname, rownum, coulumnnum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.cell(row=rownum, column=coulumnnum).value)


def writeData(path, sheetname, rownum, coulumnnum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=coulumnnum).value = data
    workbook.save(path)
